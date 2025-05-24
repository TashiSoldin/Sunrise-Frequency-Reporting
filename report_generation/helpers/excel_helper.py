from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Border, Side, PatternFill
import pandas as pd


class ExcelHelper:
    @staticmethod
    def update_template_placeholders(workbook: Workbook, replacements: dict) -> None:
        for sheet in workbook.sheetnames:
            ws = workbook[sheet]
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value in replacements:
                        cell.value = replacements[cell.value]

    @staticmethod
    def append_df_to_sheet(
        worksheet: Worksheet, df: pd.DataFrame, start_row: int
    ) -> None:
        rows = dataframe_to_rows(df, index=False, header=True)
        for row_idx, row in enumerate(rows, start=start_row):
            for col_idx, value in enumerate(row, start=1):
                worksheet.cell(row=row_idx, column=col_idx, value=value)

    @staticmethod
    def append_df_to_sheet_with_styling(
        worksheet: Worksheet, df: pd.DataFrame, start_row: int
    ) -> None:
        thin = Side(border_style="thin", color="000000")
        border = Border(top=thin, left=thin, right=thin, bottom=thin)
        last_event_col_index = df.columns.get_loc("Last Event")

        color_mapping = {
            "Attempted delivery": "f7bc00",
            "Attempted Misroute": "f7bc00",
            "Chain store floor check": "fdf900",
            "Checked in at Origin Depot": "f7bc00",
            "Consignment details captured": "f7bc00",
            "Customer query floor check": "f7bc00",
            "Event Scan Blocked": "f7bc00",
            "Floor check": "f7bc00",
            "Floor check - Booking cargo": "fdf900",
            "Floor check - Depot collection": "eac7e6",
            "Floor check - Query": "f7bc00",
            "Inbound Manifest": "f7bc00",
            "Loaded for Delivery": "00aeed",
            "Manifest Transferred": "f7bc00",
            "Mis-routed": "f7bc00",
            "Outbound Manifest Load": "fdf900",
            "POD Details Captured": "92d050",
            "POD Image Scanned": "92d050",
            "Preload": "fdf900",
            "Received at origin depot": "f7bc00",
            "Remove from manifest/tripsheet": "f7bc00",
            "Return to Client": "f7bc00",
            "Return to Depot": "f7bc00",
            "Reverse logistics floor check": "f7bc00",
            "Swadded": "f7bc00",
            "Transfer to manifest/tripsheet": "f7bc00",
            "Unload manifest/tripsheet": "f7bc00",
            "Other": "FFFFFF",
        }

        for r_idx, row in enumerate(
            dataframe_to_rows(df, index=False, header=True), start=start_row
        ):
            for c_idx, value in enumerate(row, 1):
                cell = worksheet.cell(row=r_idx, column=c_idx, value=value)
                if r_idx == start_row:
                    cell.font = Font(bold=True)
                    cell.border = border

            if r_idx > start_row:
                last_event = row[last_event_col_index]
                color = color_mapping.get(last_event, "FFFFFF")
                for c_idx in range(1, len(row) + 1):
                    cell_to_style = worksheet.cell(row=r_idx, column=c_idx)
                    cell_to_style.fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )

    @staticmethod
    def autofit_workbook_columns(workbook: Workbook) -> None:
        """
        Autofit all columns in all sheets of the given openpyxl Workbook.
        """
        for ws in workbook.worksheets:
            for col in ws.columns:
                max_length = 0
                col_letter = get_column_letter(col[0].column)
                for cell in col:
                    try:
                        cell_length = (
                            len(str(cell.value)) if cell.value is not None else 0
                        )
                        if cell_length > max_length:
                            max_length = cell_length
                    except Exception:
                        pass
                ws.column_dimensions[col_letter].width = max_length + 2

    @staticmethod
    def autofit_excel_file(file_path: str) -> None:
        """
        Open an Excel file, autofit all columns in all sheets, and save it.
        """
        from openpyxl import load_workbook

        wb = load_workbook(file_path)
        ExcelHelper.autofit_workbook_columns(wb)
        wb.save(file_path)
