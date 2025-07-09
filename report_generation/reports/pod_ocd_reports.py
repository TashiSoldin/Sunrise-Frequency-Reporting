from openpyxl import load_workbook
import pandas as pd
from tqdm import tqdm

from helpers.excel_helper import ExcelHelper
from helpers.datetime_helper import DatetimeHelper
from helpers.os_helper import OSHelper

HUB_GROUP_MAPPING_DEFINITION = {
    "JNB-HUB": {
        "dest_hubs": ["JNB", "PRY", "BFN", "WDH", "POL", "JB1"],
        "email": "steven@sunriselogistics.net",
    },
    "DUR-HUB": {
        "dest_hubs": ["DUR", "PMB", "DB1", "KZ1"],
        "email": "nishalan@sunriselogistics.net",
    },
    "CPT-HUB": {
        "dest_hubs": ["CPT"],
        "email": "shawn@sunriselogistics.net",
    },
}


class PodOcdReports:
    def __init__(self, data: dict, output_file_path: str) -> None:
        self.df: pd.DataFrame = data.get("content")
        self.hub_group_mapping = self._create_hub_to_hub_group_mapping()
        self.output_file_path = output_file_path

    def _create_hub_to_hub_group_mapping(self) -> dict:
        """
        Creates a mapping from individual destination hubs to their corresponding hub groups.

        This method transforms the hierarchical HUB_GROUP_MAPPING_DEFINITION into a flat
        dictionary for efficient lookups during report generation.

        Returns:
            dict: A dictionary mapping individual hub codes to their group names

        Example:
            If HUB_GROUP_MAPPING_DEFINITION is:
            {
                "JNB-HUB": {
                    "dest_hubs": ["JNB", "PRY", "BFN", "WDH", "POL", "JB1"],
                    "emails": [],
                },
                "DUR-HUB": {
                    "dest_hubs": ["DUR", "PMB", "DB1", "KZ1"],
                    "emails": [],
                }
            }

            The resulting mapping will be:
            {
                "JNB": "JNB-HUB",
                "PRY": "JNB-HUB",
                "BFN": "JNB-HUB",
                "WDH": "JNB-HUB",
                "POL": "JNB-HUB",
                "JB1": "JNB-HUB",
                "DUR": "DUR-HUB",
                "PMB": "DUR-HUB",
                "DB1": "DUR-HUB",
                "KZ1": "DUR-HUB",

            }
        """
        hub_group_mapping = {}
        for group_name, group_data in HUB_GROUP_MAPPING_DEFINITION.items():
            for dest_hub in group_data["dest_hubs"]:
                hub_group_mapping[dest_hub] = group_name
        return hub_group_mapping

    def _get_hub_group(self, dest_hub: str) -> str:
        return self.hub_group_mapping.get(dest_hub, dest_hub)

    def _get_hub_group_email(self, hub_group: str) -> str:
        return HUB_GROUP_MAPPING_DEFINITION.get(hub_group, {}).get("email", "")

    def sort_df(self, df: pd.DataFrame) -> pd.DataFrame:
        return df.sort_values(by="Waybill Date", ascending=True)

    def generate_report(self) -> dict:
        df = self.sort_df(self.df)
        summary = {}

        sheet_config = {
            "Verbal Missing": (df["POD Date"].isna())
            & (df["POD Image Present"] == "N"),
            "Image Missing": (df["POD Date"].notna())
            & (df["POD Image Present"] == "N"),
            "Verbal Not Captured": (df["POD Date"].isna())
            & (df["POD Image Present"] == "Y"),
        }

        df["Hub Group"] = df["Dest Hub"].map(self._get_hub_group)
        for hub_group, hub_df in tqdm(
            df.groupby("Hub Group"), desc="Generating pod ocd reports"
        ):
            template_path = OSHelper.load_template(
                "assets/", "pod_report_template.xlsx"
            )
            wb = load_workbook(template_path, data_only=False)

            # Process each sheet with the same logic
            for sheet_name, config in sheet_config.items():
                ws = wb[sheet_name]

                # Filter data based on sheet configuration
                df_filtered = hub_df[config[hub_df.index]]

                # Apply replacements
                replacements = {
                    "agent_name": hub_group,
                    "date_time": DatetimeHelper.get_precise_current_datetime(),
                    "num_missing": len(df_filtered),
                }
                ExcelHelper.update_template_placeholders_sheet(ws, replacements)

                # Append data if not empty
                if not df_filtered.empty:
                    ExcelHelper.append_df_to_sheet(ws, df_filtered, ws.max_row + 2)
                    ExcelHelper.autofit_worksheet_columns(ws)

            file_path = f"{self.output_file_path}/{hub_group}-{DatetimeHelper.get_current_datetime()}.xlsx"
            wb.save(file_path)

            summary[hub_group] = {
                "file_path": file_path,
                "client_name": hub_group,
                "email": self._get_hub_group_email(hub_group),
            }

        return summary
