from openpyxl import Workbook, load_workbook
import pandas as pd
from helpers.datetime_helper import DatetimeHelper
from helpers.os_helper import OSHelper
from helpers.dataframe_helper import DataFrameHelper
from helpers.excel_helper import ExcelHelper


class SummaryReports:
    def __init__(self, input_file_paths: list[str], output_file_path: str) -> None:
        self.input_file_paths = input_file_paths
        self.output_file_path = output_file_path

    def _summarise_file(
        self, file_name: str, file_path: str, workbook: Workbook
    ) -> dict:
        row_dict = {"File Name": file_name[:-22]}

        for sheet_name in workbook.sheetnames:
            df = DataFrameHelper.read_sheet_safely(file_path, sheet_name)

            if df.empty:
                row_dict[f"{sheet_name} Previous Month"] = 0
                row_dict[f"{sheet_name} Current Month"] = 0
                continue

            current_month = DatetimeHelper.get_today().month

            df["Month"] = pd.to_datetime(df["Waybill Date"]).dt.month
            current_month_rows = df[df["Month"] == current_month].shape[0]
            previous_months_rows = df[df["Month"] != current_month].shape[0]

            row_dict[f"{sheet_name} Previous Month"] = previous_months_rows
            row_dict[f"{sheet_name} Current Month"] = current_month_rows

        return row_dict

    def generate_report(self) -> dict:
        all_rows = []

        for input_file_path in self.input_file_paths:
            file_names = OSHelper.get_files_in_directory(input_file_path)

            for file_name in file_names:
                full_file_path = OSHelper.join_path(input_file_path, file_name)
                workbook = load_workbook(full_file_path)

                row_dict = self._summarise_file(file_name, full_file_path, workbook)
                all_rows.append(row_dict)

        df = pd.DataFrame(all_rows)
        df = DataFrameHelper.add_total_columns_for_summary(df)
        df = DataFrameHelper.add_total_row(df)

        file_path = f"{self.output_file_path}/pod-summary-report-{DatetimeHelper.get_current_datetime()}.xlsx"
        df.to_excel(file_path, index=False)

        ExcelHelper.autofit_excel_file(file_path)

        return {
            "internal": {
                "file_path": file_path,
                "client_name": "Internal",
                "email": None,
            }
        }
