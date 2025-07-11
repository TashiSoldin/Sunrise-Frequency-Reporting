from datetime import timedelta
from openpyxl import load_workbook
import pandas as pd
from tqdm import tqdm
from enums.frequency_report_enums import LastEventTypes
from helpers.excel_helper import ExcelHelper
from helpers.os_helper import OSHelper
from helpers.datetime_helper import DatetimeHelper


class FrequencyReports:
    def __init__(self, data: dict, output_file_path: str) -> None:
        self.df: pd.DataFrame = data.get("content")
        self.contact_details: pd.DataFrame = data.get("contact_email")
        self.contact_email_mapping: dict = self._get_contact_email_mapping()
        self.output_file_path = output_file_path

    def _get_contact_email_mapping(self) -> dict:
        return pd.Series(
            self.contact_details["EMAIL"].values,
            index=self.contact_details["ACCNUM"],
        ).to_dict()

    def sort_df(self, df: pd.DataFrame) -> pd.DataFrame:
        df["Last Event"] = pd.Categorical(
            df["Last Event"],
            categories=LastEventTypes.get_ordered_values(),
            ordered=True,
        )
        return df.sort_values(
            by=["Last Event", "Waybill Date"], ascending=[True, False]
        )

    def generate_report(self) -> dict:
        df = self.sort_df(self.df)
        summary = {}

        for account in tqdm(
            df["Account"].unique(), desc="Generating frequency reports"
        ):
            df_account = df[df["Account"] == account]

            completed_pod_events = ["POD Details Captured", "POD Image Scanned"]
            completed_mask = df_account["POD Date"].notna() | df_account[
                "Last Event"
            ].isin(completed_pod_events)
            completed_deliveries_df = df_account[completed_mask]
            current_deliveries_df = df_account[~completed_mask]

            # If current deliveries is empty and the last completed delivery is more than 7 days ago, continue
            valid_dates = completed_deliveries_df["Last Event Date"].dropna()
            max_last_event_date = valid_dates.max() if not valid_dates.empty else pd.NaT
            if current_deliveries_df.empty and (
                max_last_event_date is pd.NaT
                or max_last_event_date < DatetimeHelper.get_today() - timedelta(days=7)
            ):
                continue

            template_path = OSHelper.load_template(
                "assets/", "frequency_report_template.xlsx"
            )
            wb = load_workbook(template_path, data_only=False)

            replacements = {
                "client_name": df_account["Customer"].iloc[0],
                "date_time": DatetimeHelper.get_precise_current_datetime(),
            }
            ExcelHelper.update_template_placeholders(wb, replacements)

            for sheet_name, sheet_df in [
                ("Current deliveries", current_deliveries_df),
                ("Completed deliveries", completed_deliveries_df),
            ]:
                ExcelHelper.append_df_to_sheet_with_styling(
                    wb[sheet_name], sheet_df, wb[sheet_name].max_row + 2
                )

            ExcelHelper.autofit_workbook_columns(wb)

            file_path = f"{self.output_file_path}/{account}-{DatetimeHelper.get_current_datetime()}.xlsx"
            wb.save(file_path)

            summary[account] = {
                "file_path": file_path,
                "client_name": df_account["Customer"].iloc[0],
                "email": self.contact_email_mapping.get(account),
            }

        return summary
