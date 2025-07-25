from openpyxl import load_workbook
import pandas as pd
from tqdm import tqdm

from helpers.excel_helper import ExcelHelper
from helpers.datetime_helper import DatetimeHelper
from helpers.os_helper import OSHelper


class PodAgentReports:
    def __init__(self, data: dict, output_file_path: str) -> None:
        self.df: pd.DataFrame = data.get("content")
        self.agent_email: pd.DataFrame = data.get("agent_email")
        self.agent_email_mapping: dict = self._get_agent_email_mapping()
        self.output_file_path = output_file_path

    def _get_agent_email_mapping(self) -> dict:
        return pd.Series(
            self.agent_email["EMAIL"].values,
            index=self.agent_email["NAME"],
        ).to_dict()

    def sort_df(self, df: pd.DataFrame) -> pd.DataFrame:
        return df.sort_values(by="Waybill Date", ascending=True)

    def generate_report(self) -> dict:
        df = self.sort_df(self.df)
        summary = {}

        sheet_config = {
            "Verbal Missing": (df["POD Date"].isna())
            & (df["POD Image Present"] == "N"),
            "Image Missing": (df["POD Date"].notna())
            & (df["POD Image Present"] == "N"),
            "Verbal Not Captured": (df["POD Date"].isna())
            & (df["POD Image Present"] == "Y"),
        }

        for delivery_agent in tqdm(
            df["Delivery Agent"].unique(),
            desc="Generating pod agent reports",
        ):
            df_agent = df[df["Delivery Agent"] == delivery_agent]

            template_path = OSHelper.load_template(
                "assets/", "pod_report_template.xlsx"
            )
            wb = load_workbook(template_path, data_only=False)

            for sheet_name, config in sheet_config.items():
                ws = wb[sheet_name]

                # Filter data based on sheet configuration
                df_filtered = df_agent[config[df_agent.index]]

                # Apply replacements
                replacements = {
                    "agent_name": delivery_agent,
                    "date_time": DatetimeHelper.get_precise_current_datetime(),
                    "num_missing": len(df_filtered),
                }
                ExcelHelper.update_template_placeholders_sheet(ws, replacements)

                if not df_filtered.empty:
                    ExcelHelper.append_df_to_sheet(ws, df_filtered, ws.max_row + 2)
                    ExcelHelper.autofit_worksheet_columns(ws)

            file_path = f"{self.output_file_path}/{delivery_agent}-{DatetimeHelper.get_current_datetime()}.xlsx"
            wb.save(file_path)

            summary[delivery_agent] = {
                "file_path": file_path,
                "client_name": delivery_agent,
                "email": self.agent_email_mapping.get(delivery_agent),
            }

        return summary
